#!/usr/bin/env python3
"""Apply a keycap-tuner export back into lang_lut.xlsx (surgical sheet2.xml edit).

Reads the "Export changes" text the keycap tuner produces — one `=== code ===`
block per layout — and writes the changed key_lut cells + category-offset cells
straight into lang_lut.xlsx, modifying ONLY xl/worksheets/sheet2.xml so the
formula caches in the other sheets / sharedStrings stay intact (same surgical
approach as lang/_patch_xlsx.py, but it *edits existing cells* — set / clear —
instead of appending new language columns).

    python apply_tuner.py changes.txt
    pbpaste | python apply_tuner.py -                       # from the clipboard
    python apply_tuner.py changes.txt --qmk /path/to/qmk_firmware
    python apply_tuner.py changes.txt --dry-run             # just print the edits

Export format (exactly what the tuner's Export box emits):

    === ps-AF ===
    KC_Q base: U"\\f\\f" ARABIC_DAD       # set key_lut cell (var base/shift/altgr)
    KC_D altgr: <drop / empty>            # clear that cell
    [offset] letter shift H = 44          # set a category-offset cell

After applying, regenerate + rebuild:
    cd <qmk>/keyboards/polykybd/lang && cog -r lang_lut.c
    (then build / flash as usual)
"""
import sys, os, re, json, zipfile, shutil, tempfile, atexit, argparse
from xml.sax.saxutils import escape

# --- keycode -> key_lut row, mirroring oled_preview.ROW (kept inline so this
#     script needs only openpyxl, not PIL/gfx_font) -------------------------
LETTERS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
ROW = {f"KC_{c}": 2 + i for i, c in enumerate(LETTERS)}
for _i in range(1, 10): ROW[f"KC_{_i}"] = 27 + _i
ROW["KC_0"] = 37
ROW.update({"KC_MINUS": 43, "KC_EQUAL": 44, "KC_LBRC": 45, "KC_RBRC": 46,
            "KC_BACKSLASH": 47, "KC_NONUS_HASH": 48, "KC_SEMICOLON": 49,
            "KC_QUOTE": 50, "KC_GRAVE": 51, "KC_COMMA": 52, "KC_DOT": 53,
            "KC_SLASH": 54, "KC_NONUS_BACKSLASH": 55})
SET = {"letter": (57, 56), "num": (59, 58), "sym": (61, 60)}   # (voffset_row, hoffset_row)
VAR = {"base": 0, "small": 0, "shift": 1, "caps": 2, "altgr": 3}
HIDE = -128


def colname(n):                            # 1-based index -> A1 letters
    s = ""
    while n:
        n, r = divmod(n - 1, 26); s = chr(65 + r) + s
    return s


def col_of_ref(ref):                       # 'AB12' -> col index
    n = 0
    for ch in re.match(r'[A-Z]+', ref).group(0):
        n = n * 26 + (ord(ch) - 64)
    return n


def num_cell(ref, v): return f'<c r="{ref}"><v>{v}</v></c>'
def str_cell(ref, v): return f'<c r="{ref}" t="inlineStr"><is><t xml:space="preserve">{escape(v)}</t></is></c>'


# A cell is either self-closing `<c .../>` or `<c ...>...</c>`; try self-closing first.
CELL_RE = re.compile(r'<c (?:[^>]*?/>|[^>]*?>.*?</c>)', re.S)


def parse_export(text):
    """text -> { lang_code: [ (row, col_in_group, value), ... ] }.

    value is the inline-string token, an int (offset), or None to clear a cell.
    col_in_group is the 0-based offset from the language's base column.
    """
    edits = {}
    lang = None
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        m = re.match(r'^===\s*(\S+)\s*===$', line)
        if m:
            lang = m.group(1); edits.setdefault(lang, []); continue
        mo = re.match(r'^\[offset\]\s+(letter|num|sym)\s+(small|shift|altgr)\s+([HV])\s*=\s*(-?\d+)$', line)
        if mo:
            if lang is None: sys.exit("offset line before any === code === header")
            cat, var, ax, val = mo.group(1), mo.group(2), mo.group(3), int(mo.group(4))
            vrow, hrow = SET[cat]
            row = hrow if ax == "H" else vrow
            edits[lang].append((row, VAR[var], "HIDE" if val == HIDE else val))
            continue
        mk = re.match(r'^(KC_\w+)\s+(base|shift|altgr):\s*(.*)$', line)
        if mk:
            if lang is None: sys.exit("key line before any === code === header")
            kc, el, val = mk.group(1), mk.group(2), mk.group(3).strip()
            if kc not in ROW: sys.exit(f"unknown keycode {kc}")
            cell = None if val.startswith("<drop") or val in ("", "<empty>") else val
            edits[lang].append((ROW[kc], VAR[el], cell))
            continue
        sys.exit(f"unparseable export line: {raw!r}")
    return edits


def set_cell(xml, r, c, new_cell):
    """Insert / replace / delete (new_cell=None) the cell at (row r, col c) in sheet2.xml."""
    ref = colname(c) + str(r)
    rm = re.search(r'(<row r="%d"[^>]*>)(.*?)(</row>)' % r, xml, re.S)
    if not rm:                              # row absent — every populated row exists, so this is unexpected
        if new_cell is None:
            return xml
        sys.exit(f"row {r} not found in sheet2.xml (cannot place {ref})")
    head, body, tail = rm.group(1), rm.group(2), rm.group(3)
    existing = next((cm for cm in CELL_RE.finditer(body)
                     if re.search(r'r="([A-Z]+\d+)"', cm.group(0)).group(1) == ref), None)
    if existing:
        body = body[:existing.start()] + (new_cell or "") + body[existing.end():]
    elif new_cell is not None:             # insert before the first cell of a higher column
        pos = len(body)
        for cm in CELL_RE.finditer(body):
            if col_of_ref(re.search(r'r="([A-Z]+\d+)"', cm.group(0)).group(1)) > c:
                pos = cm.start(); break
        body = body[:pos] + new_cell + body[pos:]
    return xml[:rm.start()] + head + body + tail + xml[rm.end():]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('export', help="the tuner's exported text, or - for stdin")
    _here = os.path.dirname(os.path.abspath(__file__))
    ap.add_argument('--qmk', default=os.path.join(os.path.dirname(os.path.dirname(_here)), 'qmk_firmware'))
    ap.add_argument('--xlsx', default=None, help='override path to lang_lut.xlsx')
    ap.add_argument('--dry-run', action='store_true', help='print the resolved edits, write nothing')
    a = ap.parse_args()

    text = sys.stdin.read() if a.export == '-' else open(a.export, encoding='utf-8').read()
    edits = parse_export(text)
    if not edits or not any(edits.values()):
        print("no edits in export — nothing to do"); return

    XLSX = a.xlsx or os.path.join(a.qmk, 'keyboards', 'handwired', 'polykybd', 'lang', 'lang_lut.xlsx')
    if not os.path.exists(XLSX):
        sys.exit(f"lang_lut.xlsx not found at {XLSX} (use --qmk or --xlsx)")

    from openpyxl import load_workbook
    wb = load_workbook(XLSX, data_only=True, read_only=True)
    sh = wb['key_lut']
    langs, i = [], 0
    while sh.cell(row=1, column=2 + i * 4).value:
        langs.append(sh.cell(row=1, column=2 + i * 4).value); i += 1
    wb.close()

    # resolve every edit to an absolute (row, col, cell_xml|None)
    resolved, total = [], 0
    for lang, items in edits.items():
        if not items:
            continue
        if lang not in langs:
            sys.exit(f"layout {lang!r} is not in lang_lut.xlsx (have {len(langs)} layouts)")
        base = 2 + langs.index(lang) * 4
        for row, var, val in items:
            c = base + var
            ref = colname(c) + str(row)
            if val is None:
                cell = None; shown = "<clear>"
            elif isinstance(val, int):
                cell = num_cell(ref, val); shown = str(val)
            else:
                cell = str_cell(ref, val); shown = val
            resolved.append((row, c, cell))
            total += 1
            print(f"  {lang:8s} {ref:>5s}  {shown}")
    print(f"{total} cell edit(s) across {sum(1 for v in edits.values() if v)} layout(s)")

    if a.dry_run:
        print("(dry run — nothing written)"); return

    tmp = tempfile.mkdtemp(prefix="xlsx_apply_")
    atexit.register(shutil.rmtree, tmp, ignore_errors=True)
    with zipfile.ZipFile(XLSX) as z:
        z.extractall(tmp)
    sheet2 = os.path.join(tmp, "xl/worksheets/sheet2.xml")
    xml = open(sheet2, encoding="utf-8").read()
    for row, c, cell in resolved:
        xml = set_cell(xml, row, c, cell)
    open(sheet2, "w", encoding="utf-8").write(xml)

    out = XLSX + ".new"
    with zipfile.ZipFile(XLSX) as zin, zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = open(os.path.join(tmp, item.filename), "rb").read() if item.filename == "xl/worksheets/sheet2.xml" else zin.read(item.filename)
            zout.writestr(item, data)
    shutil.move(out, XLSX)
    print(f"patched {XLSX}")
    print("next:  cd %s/keyboards/polykybd/lang && cog -r lang_lut.c" % a.qmk)


if __name__ == '__main__':
    main()
