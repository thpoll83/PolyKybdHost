#!/usr/bin/env python3
"""Render a pixel-exact preview of a single PolyKybd keycap OLED (72x40) for a
chosen language layout - the "rendering analysis" companion to emoji_demo.py.

It reuses gfx_font.load_all_fonts() (the same pixel-exact GFX renderer
emoji_demo uses in --font-mode gfx) and replicates the firmware's per-key draw
in split72/keymaps/default/keymap.c: translate_keycode() picks the base glyph
(with the NULL->en-US fallback), get_setting() supplies the letter/num/sym
h/v offsets, and the unshifted view also lays out the shift-preview (with the
same clear/clamp/stagger rules) and the AltGr-preview. Glyphs, sizes and
positions therefore match what the hardware draws.

Data comes straight from lang/lang_lut.xlsx + lang/named_glyphs.h, so it always
reflects the current spreadsheet (run it right after editing, before flashing).

Usage:
    python tools/oled_preview.py --lang ka-GE                 # contact sheet of all keys
    python tools/oled_preview.py --lang ta-IN --key KC_Q      # one key, big
    python tools/oled_preview.py --lang vi-VN --out /tmp/vn.png --cell-scale 4
    python tools/oled_preview.py --lang hy-AM --shift          # show the shifted view
"""
from __future__ import annotations
import argparse, os, re, sys
from PIL import Image, ImageDraw, ImageFont

from gfx_font import (load_all_fonts, load_ui_font, MID_FONT_FILE, MID_FONT_SYMBOL,
                      OLED_W, OLED_H, BUFFER_X, BASELINE)

HERE = os.path.dirname(os.path.abspath(__file__))
HOST_REPO = os.path.dirname(HERE)
HOME = os.path.dirname(HOST_REPO)
HIDE = -128
SCREEN_WIDTH = 72
# How many px outside the real 72x40 viewport to KEEP and render (instead of
# silently clipping like the hardware does). oled_to_rgb paints this border red
# with any lit pixels in yellow, so glyphs that get cut off on the device are
# obvious in the preview. 0 = hardware-exact (no margin). Overridable via --overshoot.
OVERSHOOT = 2

# ---- display-list ops (base/disp_array.c gfx_text_run / base/font_lookup.c) ----
# A legend is a mini display list, not just text: codepoints below 0x20 are ops, and
# some of them consume the codepoints that follow as ARGUMENTS. Renderer interprets
# the cursor nudges and the two size ops; everything else it can only SKIP, because
# drawing it would need a primitive this model does not have (a rounded rect, a
# rotated glyph, an absolute buffer position). Skipping is still much better than the
# old fall-through, which measured and drew the op byte AND each of its arguments as
# a substituted '!'.
HINT_SMALL = 0x10      # rest of the run at half scale (kdisp_write_gfx_char_half)
ALTGR_HALF_MIN_INK_H = 7   # halve the AltGr hint only when its ink is taller than this
HINT_MID = 0x16        # rest of the run from the standalone 19px UI face
CURSOR_OPS = frozenset({0x05, 0x06, 0x08, 0x09, 0x0A, 0x0B, 0x0C, 0x0D, 0x18})
SUPPORTED_OPS = CURSOR_OPS | {HINT_SMALL, HINT_MID,
                              0x0E,   # MOVE  - absolute buffer position
                              0x13,   # BADGE - the lock-indicator box
                              0x14,   # ERASE - plot as a hole
                              0x15}   # ROT   - rotated, halved glyph
# op -> how many following codepoints are its arguments.
OP_ARGS = {0x0E: 2,    # MOVE (x, y)  - an ABSOLUTE buffer position
           0x0F: 1,    # HALF    (glyph) - composite at the cursor, no advance
           0x11: 1,    # THIN    (glyph)
           0x12: 2,    # FRAME   (w, h)
           0x13: 3,    # BADGE   (w, h, style)
           0x14: 0,    # ERASE   - a plotter mode, no extent
           0x15: 2}    # ROT     (angle, glyph)


def _half_floor(v: int) -> int:
    """Floor division by two — glyph_half_floor() in base/font_lookup.h.

    Glyph offsets are negative (above the baseline) and C truncates toward zero,
    which rounds those the wrong way and puts a lowercase glyph 1px off its run's
    baseline. Python's // already floors; this exists to name the rule.
    """
    return v // 2


def _trunc_div(a: int, b: int) -> int:
    """C integer division — truncate toward ZERO, not Python's floor.

    ⚠️ The \\v and \\t steps are `x += (x / N + 1) * N` on a cursor that can be
    NEGATIVE relative to the origin: MID_TWO_LINE lifts the first baseline 10px
    before stepping down a line. Python's `//` floors, so (-10)//15 is -1 and the
    step came out as ZERO — the second line landed on top of the first, which is
    the exact collision the size ops were added to avoid.
    """
    q = abs(a) // abs(b)
    return q if (a >= 0) == (b >= 0) else -q


def _walk_ops(cps):
    """Yield ``(is_op, cp)`` for a display list, consuming each op's ARGUMENTS once.

    The argument-skipping rule is the part of this that has been got wrong before —
    a coordinate byte that is itself an op byte (``HINT_SZ_STOPSQ`` is (15,15), i.e.
    two HALFs) silently latching a font for the rest of the run. So it lives in ONE
    walk that measuring, drawing and the can-I-draw-this check all share, rather
    than in three copies that have to be kept in step.

    ⚠️ A TRUNCATED op does not skip: the firmware guards each ``text += n`` on the
    arguments actually being there, so the op is consumed and the stray bytes are
    then read as ordinary codepoints.
    """
    skip = 0
    for i, cp in enumerate(cps):
        if skip:
            skip -= 1
            continue
        if cp in OP_ARGS:
            n = OP_ARGS[cp]
            args = ()
            if n and len(cps) - i - 1 >= n:
                skip = n
                args = tuple(cps[i + 1:i + 1 + n])
            yield True, cp, args
            continue
        yield cp < 0x20, cp, ()


# ---- composite-op primitives, ported from the firmware --------------------------
# base/disp_array.c (kdisp_draw_badge_rect / rr_row_inset /
# kdisp_draw_glyph_rot_half_at) and base/font_lookup.c (kdisp_gfx_rot_half_extent).
# Ported rather than approximated because these draw the LOCK BADGES and the context
# -menu pointer, and a keycap that is nearly right is the failure this whole module
# exists to avoid.
KDISP_BADGE_RADIUS = 2
KDISP_BADGE_BORDER = 2      # the released badge's stroke, matching the baked glyphs

# sin(15 deg * i) in 8.8 fixed point, i = 0..23 — the firmware's s_sin15 verbatim.
_SIN15 = (0, 66, 128, 181, 222, 247, 256, 247,
          222, 181, 128, 66, 0, -66, -128, -181,
          -222, -247, -256, -247, -222, -181, -128, -66)


def _rr_row_inset(j, top, bot, r):
    """How far row `j` is inset by a corner radius `r` — the firmware's rr_row_inset.

    ⚠️ This scanline formula, NOT a Bresenham arc: the two disagree about what
    "r = 2" looks like (Bresenham insets 1,0 where this gives 2,1,0), and the badge
    has to match the baked ICON_CAPSLOCK_* corners, which inset 2,1,0.
    """
    d = (top + r - j) if j < top + r else ((j - (bot - r)) if j > bot - r else 0)
    if d <= 0:
        return 0
    rem = r * r - d * d
    k = 0
    while (k + 1) * (k + 1) <= rem:
        k += 1
    return r - k


def draw_badge_rect(plot, x, y, width, height, r, border):
    """A rounded lock badge: `border` 0 fills it solid, else strokes a ring that thick.

    ⚠️ The hole is NOT a true concentric offset. That would imply an inner radius of
    `r - border`, which at `r == border` is a perfectly square inner corner — one
    pixel short of the baked ICON_CAPSLOCK_OFF, whose hole still insets 1 on its
    first row. Reported from hardware as "it misses a single pixel on the inside
    corner", so the radius is floored at 1 whenever the outer corner is rounded.
    """
    if width < 2 or height < 2:
        return
    r = max(0, min(r, (width - 1) // 2, (height - 1) // 2))
    border = max(0, border)
    x0, y0 = x, y
    x1, y1 = x + width - 1, y + height - 1
    hx0, hy0, hx1, hy1 = x0 + border, y0 + border, x1 - border, y1 - border
    hr = r - border
    if hr < 1:
        hr = 1 if r > 0 else 0
    for j in range(y0, y1 + 1):
        ins = _rr_row_inset(j, y0, y1, r)
        a, b = x0 + ins, x1 - ins
        if border > 0 and hy0 <= j <= hy1 and hx0 <= hx1:
            hins = _rr_row_inset(j, hy0, hy1, hr)
            ha, hb = hx0 + hins, hx1 - hins
            for i in range(a, b + 1):
                if i < ha or i > hb:
                    plot(i, j)
            continue
        for i in range(a, b + 1):
            plot(i, j)


def rot_half_extent(w, h, step):
    """The rotated-and-halved frame for a glyph — kdisp_gfx_rot_half_extent.

    Returns `(ct, st, cx, cy, x0, y0, out_w, out_h)` in the firmware's 8.8 fixed
    point. It lives beside the bbox interpreter in the firmware for the same reason
    it is one function here: the measure and the draw must agree about exactly what
    is plotted, or a MOVE'd mark is clamped against the wrong box.
    """
    # Screen y runs DOWN, so a visually counter-clockwise turn is a NEGATIVE angle in
    # this arithmetic — hence the (24 - step) index. The op's argument is stated in
    # the direction a reader means by "counter-clockwise", not the sign the maths uses.
    idx = (24 - (step % 24)) % 24
    ct = _SIN15[(idx + 6) % 24]
    st = _SIN15[idx]
    cx = ((w - 1) << 8) // 2
    cy = ((h - 1) << 8) // 2
    xs, ys = [], []
    for c in range(4):
        sx = cx if (c & 1) else -cx
        sy = cy if (c & 2) else -cy
        xs.append(_asr8(sx * ct - sy * st))
        ys.append(_asr8(sx * st + sy * ct))
    x0, x1 = min(xs), max(xs)
    y0, y1 = min(ys), max(ys)
    return (ct, st, cx, cy, x0, y0,
            ((_asr8(x1 - x0) + 1) + 1) // 2,
            ((_asr8(y1 - y0) + 1) + 1) // 2)


def _int8(v):
    """A codepoint argument read as the firmware reads it: `(int8_t)text[n]`.

    The coordinates are written as `U"\\x48"`-style bytes, so anything above 127 is
    a NEGATIVE position. Reading them unsigned puts the mark off the panel.
    """
    v &= 0xFF
    return v - 256 if v > 127 else v


def _asr8(v):
    """`v >> 8` with C's arithmetic shift on a negative value.

    Python's `>>` already floors, which is what an arithmetic shift does — this
    exists to say so, because `//` would be wrong for the // 2 divisions above and
    it is easy to reach for the same operator twice.
    """
    return v >> 8


def load_renderer(font_dir: str) -> "Renderer":
    """A Renderer with the ALL_FONTS pool AND the HINT_MID face.

    Prefer this over ``Renderer(load_all_fonts(d))``: without the mid face a legend
    carrying \\x16 renders at full size, which for the settings keys stacks two lines
    of text on top of each other.
    """
    # ⚠️ Degrade rather than raise if the mid header is missing. It is a SECOND
    # prerequisite, and the caller's other legends do not need it -- taking the whole
    # renderer down over it is the "load the two halves independently" mistake that
    # once left a machine previewing nothing but macros. Without it \x16 is reported
    # by unsupported_ops(), so those few legends are refused instead of drawn wrong.
    try:
        mid = [load_ui_font(font_dir, MID_FONT_FILE, MID_FONT_SYMBOL)]
    except Exception:
        mid = None
    return Renderer(load_all_fonts(font_dir), mid_fonts=mid)

# keycode -> lang_lut row, in translate_keycode order
LETTERS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
ROW = {f"KC_{c}": 2 + i for i, c in enumerate(LETTERS)}
for i in range(1, 10): ROW[f"KC_{i}"] = 27 + i
ROW["KC_0"] = 37
ROW.update({"KC_MINUS": 43, "KC_EQUAL": 44, "KC_LBRC": 45, "KC_RBRC": 46,
            "KC_BACKSLASH": 47, "KC_NONUS_HASH": 48, "KC_SEMICOLON": 49,
            "KC_QUOTE": 50, "KC_GRAVE": 51, "KC_COMMA": 52, "KC_DOT": 53,
            "KC_SLASH": 54, "KC_NONUS_BACKSLASH": 55})
# contact-sheet layout (rows of keycodes); None = gap
SHEET = [
    ["KC_GRAVE","KC_1","KC_2","KC_3","KC_4","KC_5","KC_6","KC_7","KC_8","KC_9","KC_0","KC_MINUS","KC_EQUAL"],
    ["KC_Q","KC_W","KC_E","KC_R","KC_T","KC_Y","KC_U","KC_I","KC_O","KC_P","KC_LBRC","KC_RBRC","KC_BACKSLASH"],
    ["KC_A","KC_S","KC_D","KC_F","KC_G","KC_H","KC_J","KC_K","KC_L","KC_SEMICOLON","KC_QUOTE","KC_NONUS_HASH"],
    ["KC_NONUS_BACKSLASH","KC_Z","KC_X","KC_C","KC_V","KC_B","KC_N","KC_M","KC_COMMA","KC_DOT","KC_SLASH"],
]
VAR_SMALL, VAR_SHIFT, VAR_CAPS, VAR_ALTGR = 0, 1, 2, 3

# Keycap legend SIZE (HID cmd 34, firmware protocol v13+). Mirrors
# glyph_size_base[] / glyph_size_baseline[] in poly_keymap.c — the bigger faces are
# the same characters emitted at an offset into supplementary private-use plane 15,
# so the resident small face cannot win the front-to-back lookup for them.
GLYPH_SIZE_BASE     = {0: 0, 1: 0xF0000, 2: 0xF3000}
GLYPH_SIZE_BASELINE = {0: 21, 1: 25, 2: 28}   # [0] unused: size S keeps the language offset
GLYPH_SIZE_MAX_LEN  = 4                        # glyph_size_remap()'s length guard
GLYPH_SIZE_NAMES    = {0: "small", 1: "medium", 2: "large"}
SET = {"letter": (57, 56), "num": (59, 58), "sym": (61, 60)}   # (voffset_row, hoffset_row)
# {letter.altgrhalf} -- per-LAYOUT opt-in to the half-size AltGr hint, read from
# the VAR_ALTGR sub-column. A settings row like the six above, so `Lang` needs no
# special case; only this row number and the gate in render_key() know about it.
ALTGR_HALF_ROW = 62


# ---- named glyphs + cell resolution ---------------------------------------
def parse_u_string(content: str) -> list[int]:
    """Codepoints from the body of a C U"..."/u"..." literal (handles \\xHH.., \\f \\v \\b \\n \\r \\t \\x05 \\x18)."""
    cps, i = [], 0
    simple = {'f': 0x0c, 'v': 0x0b, 'b': 0x08, 'n': 0x0a, 'r': 0x0d, 't': 0x09, '0': 0, '\\': 0x5c, '"': 0x22}
    while i < len(content):
        c = content[i]
        if c == '\\' and i + 1 < len(content):
            nx = content[i + 1]
            if nx == 'x':
                j = i + 2; h = ''
                while j < len(content) and content[j] in '0123456789abcdefABCDEF':
                    h += content[j]; j += 1
                cps.append(int(h, 16)); i = j; continue
            cps.append(simple.get(nx, ord(nx))); i += 2; continue
        cps.append(ord(c)); i += 1
    return cps


# ---- C macro reading (shared with keycap_preview, which re-exports these) --------
# These live HERE, beside the glyph loader, because the macros they expand are the
# ones `named_glyphs.h` defines. keycap_preview needs the same expander for the
# legend switch in `keycode_helper.c`, and the lower layer is the one that can be
# imported by both.


def _strip_c_comments(s: str) -> str:
    s = re.sub(r"/\*.*?\*/", "", s, flags=re.S)
    return re.sub(r"//[^\n]*", "", s)


_FUNC_MACRO_RE = re.compile(r"#define\s+(\w+)\(([^)]*)\)\s*((?:[^\n\\]|\\\n|\\)*)")


def parse_function_macros(*texts: str) -> dict:
    """Function-like `#define NAME(a, b) body` -> {name: (params, body)}.

    The glyph loader in `oled_preview` handles only OBJECT-like macros, and says so:
    "a macro this loader has not seen falls back to drawing its own IDENTIFIER". The
    settings legends are built from function-like ones (`MID_TWO_LINE`,
    `MID_WORD_OVER_ICON`, and `SETTING_LBL` which wraps them), so three keys rendered
    the literal text `MID_T…`, `idle_s…`, `glyph…` instead of their legends.
    """
    out = {}
    for text in texts:
        for m in _FUNC_MACRO_RE.finditer(_strip_c_comments(text)):
            params = [x.strip() for x in m.group(2).split(",") if x.strip()]
            # ⚠️ Strip LINE CONTINUATIONS only. A blanket backslash strip also eats the
            # escapes inside the literals -- `U"\\f\\f\\f"` became `U" f f f"` and the
            # legend rendered the letter f four times over.
            #
            # ⚠️ ...and a C continuation is ONE backslash. This pattern asked for TWO
            # (`\\\\` in a raw string), so it never matched and every multi-line legend
            # macro kept a literal `\\` at the front -- which resolves to a real glyph, so
            # the five settings keycaps drew a backslash before their label. It was
            # invisible while those legends were refused for using HINT_MID.
            body = re.sub(r"\\\s*\n\s*", " ", m.group(3)).strip()
            if params and body:
                out[m.group(1)] = (params, body)
    return out


def expand_function_macros(expr: str, macros: dict, depth: int = 6) -> str:
    """Expand `SETTING_LBL("IDLE:", "Pulse")` down to its literals.

    ⚠️ Bounded rather than recursive-until-stable: these nest (SETTING_LBL wraps
    MID_TWO_LINE) but a macro that expanded to itself would otherwise hang the editor
    while it painted a key.
    """
    for _ in range(depth):
        m = _find_macro_call(expr, macros)
        if m is None:
            return expr
        name, args, start, end = m
        params, body = macros[name]
        if len(args) != len(params):
            return expr                      # arity mismatch: leave it alone
        for param, arg in zip(params, args):
            # ⚠️ The replacement is a FUNCTION, not a string. `re.sub` reads a string
            # replacement as a TEMPLATE, so an argument carrying a C escape --
            # `HINT_MOVE(HINT_POS_CTXPTR)` expands to `U"\x42" U"\x0C"` -- raised
            # `bad escape \x` and took the whole glyph load down with it. Harmless
            # while this only ever expanded the settings labels ("IDLE:", "Pulse"),
            # which is why it went unnoticed until the glyph macros came through here.
            paste = "U" + arg
            body = re.sub(r"U\s*##\s*\b" + re.escape(param) + r"\b",
                          lambda _m, r=paste: r, body)
            body = re.sub(r"\b" + re.escape(param) + r"\b",
                          lambda _m, r=arg: r, body)
        expr = expr[:start] + body + expr[end:]
    return expr


def _find_macro_call(expr: str, macros: dict):
    """The first `NAME(...)` in `expr` whose NAME we know, with its split arguments."""
    for m in re.finditer(r"\b(\w+)\s*\(", expr):
        if m.group(1) not in macros:
            continue
        depth, i, args, cur = 0, m.end() - 1, [], ""
        while i < len(expr):
            ch = expr[i]
            if ch == "(":
                depth += 1
                if depth == 1:
                    i += 1
                    continue
            elif ch == ")":
                depth -= 1
                if depth == 0:
                    args.append(cur.strip())
                    return m.group(1), [a for a in args if a], m.start(), i + 1
            elif ch == "," and depth == 1:
                args.append(cur.strip())
                cur = ""
                i += 1
                continue
            cur += ch
            i += 1
    return None


def load_named_glyphs(path: str, *extra: str) -> dict[str, list[int]]:
    """Named-glyph macros, merged over every header given (later files win).

    ⚠️ `lang/named_glyphs.h` is NOT the only place a keycap legend macro lives —
    `keycode_helper.h` defines the Intl ones (INTL_LAYER_LEGEND / _PICKER_ / _REMAP_),
    and `keycode_to_static_text()` returns those by name. A macro this loader has not
    seen falls back to drawing its own IDENTIFIER, so the Intl key rendered as the
    text "INTL_" on every board preview instead of İñțł. Sibling headers are picked
    up automatically below; pass more explicitly if a legend ever moves again.

    ⚠️ The body is read WHOLE — every literal, every nested macro, every
    function-like `HINT_*()` call. It used to be matched with a single-literal
    pattern that stopped at the first `U"..."`, which is worse than not matching at
    all: a truncated legend still renders, so it reads as correct-but-incomplete
    rather than as missing. Three keycaps shipped that way (2026-09-01) —
    `ICON_CONTEXT_MENU` collapsed to `U" "` and drew nothing, `ICON_SCRLOCK_OFF/ON`
    to `U"Scr"` so the lock badge vanished, and `ICON_PAUSE_TEXT` to a bare cursor
    op. The six `HINT_POS_*` / `HINT_SZ_*` constants were truncated to their x with
    the y dropped, which would have mis-placed anything that used them.
    """
    out = {}
    paths = [path, *extra]
    sibling = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(path))),
                           'keycode_helper.h')
    if not extra and os.path.exists(sibling):
        paths.append(sibling)
    texts = []
    for p in paths:
        if not os.path.exists(p):
            continue
        with open(p, encoding='utf-8') as fh:
            texts.append(fh.read())
    if not texts:
        return out

    # Function-like macros first: an object macro's body can CALL one
    # (`HINT_MOVE(HINT_POS_CTXPTR)`), so they have to be expandable before any body
    # is resolved.
    calls = parse_function_macros(*texts)
    bodies = {}
    for text in texts:
        # ⚠️ Join C line continuations before matching. Every multi-line legend macro
        # -- which is most of the interesting ones -- otherwise presents as a body
        # that ends mid-expression.
        joined = re.sub(r'\\\s*\n\s*', ' ', _strip_c_comments(text))
        for m in re.finditer(r'^[ \t]*#define[ \t]+([A-Za-z_]\w*)[ \t]+(\S.*?)[ \t]*$',
                             joined, re.M):
            bodies[m.group(1)] = m.group(2).strip()
    _resolve_macro_bodies(out, bodies, calls)
    return out


_LITERAL_OR_NAME = re.compile(r'[uU]"(?:\\.|[^"\\])*"|[A-Za-z_]\w*')


def _resolve_macro_bodies(out: dict, bodies: dict, calls: dict) -> None:
    """Expand each `#define` body down to codepoints, in `out`.

    A body is a sequence of `U"..."` literals and other macro names, possibly with
    function-like calls mixed in. Expanded in bounded passes rather than
    iterate-until-stable: they nest a level or two in practice, and a macro that
    referenced itself would otherwise spin while the editor paints a key.

    A body still holding an unknown name after the passes is LEFT OUT, so
    `unresolved_tokens` reports it and the legend is refused. That is the honest
    outcome: a partial expansion draws a keycap missing a glyph with nothing to say
    so, which is the very defect this function exists to close.
    """
    for _ in range(6):
        progressed = False
        for name, body in bodies.items():
            if name in out:
                continue
            expanded = expand_function_macros(body, calls) if calls else body
            if '(' in expanded and re.search(r'\b[A-Za-z_]\w*\s*\(', expanded):
                continue                 # an unexpanded call: try again next pass
            cps, ok = [], True
            for tok in _LITERAL_OR_NAME.finditer(expanded):
                t = tok.group(0)
                if t[0] in 'uU' and t[1:2] == '"':
                    cps += parse_u_string(t[2:-1])
                elif t in out:
                    cps += list(out[t])
                else:
                    ok = False
                    break
            if ok and cps:
                out[name] = cps
                progressed = True
        if not progressed:
            break


class Lang:
    def __init__(self, xlsx: str, named: dict):
        from openpyxl import load_workbook
        wb = load_workbook(xlsx, data_only=True, read_only=True)
        ws = wb['key_lut']
        self.named = named
        # Materialise the sheet once. A read_only worksheet streams forward, so random
        # `.cell(row, col)` access re-scans and is O(n) per call - fine for one layout,
        # but it made --all (every layout x ~49 keys) crawl. A single iter_rows pass
        # into a {(row,col): value} grid makes every later lookup O(1).
        self.grid = {}
        for r, rowvals in enumerate(ws.iter_rows(values_only=True), start=1):
            for c, v in enumerate(rowvals, start=1):
                if v is not None and v != '':
                    self.grid[(r, c)] = v
        wb.close()
        self.langs = []
        i = 0
        while self.grid.get((1, 2 + i * 4)):
            self.langs.append(self.grid[(1, 2 + i * 4)])
            i += 1

    def basecol(self, lang): return 2 + self.langs.index(lang) * 4

    def cell(self, lang_idx0, row, var):
        """raw cell value at language (0-based index), row, variation column."""
        return self.grid.get((row, 2 + lang_idx0 * 4 + var))

    def resolve(self, val) -> list[int] | None:
        """cell value -> list of codepoints (the U'...' the firmware would build), or None."""
        if val is None or val == '': return None
        if isinstance(val, (int, float)): return [ord(c) for c in str(int(val))]
        s = str(val).strip()
        # Tokenise like the firmware's make_key: a u"..."/U"..." literal is ONE token
        # even if it contains spaces (e.g. u"[ {", u"` ~"); only whitespace OUTSIDE a
        # quoted literal separates tokens (e.g. u"\f\f" MICRO_SIGN). Splitting on plain
        # whitespace shattered space-containing literals (the ro-RO/bracket-key bug).
        cps = []
        for m in re.finditer(r'[uU]"(?:\\.|[^"\\])*"|\S+', s):
            cps += self.resolve_token(m.group(0))
        return cps or None

    def unresolved_tokens(self, val) -> list:
        """The MACRO-looking tokens in `val` that this table has no glyphs for.

        ⚠️ `resolve_token` falls back to parsing an unknown token as the body of an
        implicit `U"..."`, which is right for a bare cell and WRONG for a macro name:
        the keycap then draws the literal word `ICON_MEDIA_STOP`. That shipped
        (2026-09-01), and it is invisible from the code -- the legend resolves, the
        renderer supports every op in it, and the picture is a line of text.

        So a caller that shows the result as if it were the keyboard should refuse a
        legend this reports on, exactly as it refuses one `Renderer.unsupported_ops`
        reports on. A token is macro-looking only if it is ALL-CAPS with no lowercase
        and at least two characters -- an ordinary legend body is a `U"..."` literal
        or lowercase text, so this cannot fire on real content.
        """
        if val is None or val == '' or isinstance(val, (int, float)):
            return []
        out = []
        for m in re.finditer(r'[uU]"(?:\\.|[^"\\])*"|\S+', str(val).strip()):
            t = m.group(0)
            if t in self.named or not re.fullmatch(r'[A-Z][A-Z0-9_]+', t):
                continue
            out.append(t)
        return out

    def resolve_token(self, t: str) -> list[int]:
        if t in self.named: return list(self.named[t])
        m = re.match(r'^[uU]"((?:\\.|[^"\\])*)"$', t)
        if m: return parse_u_string(m.group(1))
        # A bare cell is the body of an implicit U"..." (make_key wraps it), so it can
        # carry \x.. / \f escapes — parse them instead of rendering the literal text
        # (a bare "\xb4\xb4" is U+00B4 U+00B4, not the 8 chars backslash-x-b-4...).
        return parse_u_string(t)

    # translate_keycode (small) with NULL -> en-US(0) fallback; returns (used_lang, cps|None)
    def small(self, lang_idx, row):
        v = self.cell(lang_idx, row, VAR_SMALL)
        cps = self.resolve(v)
        if cps is None:
            return 0, self.resolve(self.cell(0, row, VAR_SMALL))
        return lang_idx, cps

    def var(self, used_lang, row, var):
        return self.resolve(self.cell(used_lang, row, var))


def get_setting(L: Lang, row: int, lang_idx: int, var: int) -> int:
    v = L.cell(lang_idx, row, var)
    if v is None or v == '': return 0
    if isinstance(v, str) and v.strip().upper() == 'HIDE': return HIDE
    return int(v)


# ---- GFX text blit (parameterised y; mirrors kdisp_write_gfx_text_cy) ------
class Renderer:
    def __init__(self, fonts, mid_fonts=None):
        self.fonts = fonts
        self.base_yadv = fonts[0].yAdvance
        # HINT_MID's face is a SEPARATE pool, not another entry in `fonts`: the
        # firmware draws it through a single-font array, so its own yAdvance is the
        # baseline reference for the glyphs it supplies (adjustment 0). None = this
        # renderer was built without it and cannot follow \x16 (see unsupported_ops).
        self.mid_fonts = list(mid_fonts) if mid_fonts else None

    @staticmethod
    def _font_in(pool, cp):
        for f in pool:
            if f.first <= cp <= f.last:
                g = f.glyphs[cp - f.first]
                # skip an empty padding gap (non-contiguous range filled with 0x0
                # glyphs) so a later font with the real glyph wins; a real space is
                # (1,1)/advance>0 and is never skipped. Mirrors disp_array.c.
                if g['width'] == 0 and g['height'] == 0 and g['xAdvance'] == 0:
                    continue
                return f
        return None

    def _font(self, cp):
        return self._font_in(self.fonts, cp)

    def _pool(self, cp, mid):
        """(pool, baseline-reference yAdvance) for one codepoint of a run.

        Mirrors the draw's PER-GLYPH mid fallback: the mid face is ASCII-only, so
        anything outside it comes from the caller's pool at full size — which is what
        makes a word-over-icon legend possible at all. The baseline reference has to
        follow that per-glyph choice, because the draw aligns each glyph by
        `font.yAdvance - fonts[0].yAdvance` and fonts[0] is the mid face only for the
        glyphs the mid face supplied.
        """
        if mid and self.mid_fonts and self._font_in(self.mid_fonts, cp) is not None:
            return self.mid_fonts, self.mid_fonts[0].yAdvance
        return self.fonts, self.base_yadv

    def _draw_glyph_rot_half(self, plot, x, y, ch, step):
        """Rotate a glyph counter-clockwise by step*15 deg, halve it, plot at (x, y).

        Mirrors kdisp_draw_glyph_rot_half_at: no baseline align, no xOffset, no
        cursor advance — the same "composite one icon into a hint" contract HALF has.

        ⚠️ It rotates at FULL resolution and halves afterwards, which is why the 2x2
        loop is INSIDE the pixel loop rather than rotating an already-halved glyph.
        Halving first throws away the very pixels the rotation needs to reconstruct
        an edge, and the arrowhead this exists for came out visibly broken that way.
        """
        f = self._font(ch)
        if f is None or not (f.first <= ch <= f.last):
            return
        g = f.glyphs[ch - f.first]
        w, h = g['width'], g['height']
        if w <= 0 or h <= 0:
            return
        cb = (h + 7) >> 3
        bo = g['bitmapOffset']
        ct, st, cx, cy, x0, y0, ow, oh = rot_half_extent(w, h, step)
        for dy in range(oh):
            for dx in range(ow):
                lit = False
                for o in range(4):
                    # The full-resolution destination pixel this quarter stands for,
                    # centre-relative so the inverse rotation is a pure rotate.
                    fx = ((dx * 2 + (o & 1)) << 8) + x0
                    fy = ((dy * 2 + (o >> 1)) << 8) + y0
                    sx = _asr8(fx * ct + fy * st) + cx
                    sy = _asr8(-fx * st + fy * ct) + cy
                    ix, iy = _asr8(sx + 128), _asr8(sy + 128)
                    if ix < 0 or ix >= w or iy < 0 or iy >= h:
                        continue
                    if f.bitmap[bo + ix * cb + (iy >> 3)] & (1 << (iy & 7)):
                        lit = True
                        break
                if lit:
                    plot(x + dx, y + dy)


    def unsupported_ops(self, cps) -> set:
        """The ops in `cps` this renderer cannot follow — empty means it can draw it.

        A caller that shows the result as if it were the keyboard should refuse a
        legend this reports on, rather than draw one that is quietly missing a frame,
        a badge or a MOVE'd mark. HINT_MID counts as unsupported when no mid face was
        loaded, since the run would then silently render at full size.
        """
        bad = {cp for is_op, cp, _a in _walk_ops(cps)
               if is_op and cp not in SUPPORTED_OPS}
        if self.mid_fonts is None and HINT_MID in cps:
            bad.add(HINT_MID)
        return bad

    def bounds(self, cps):
        """(min,max) horizontal extent relative to cursor 0 — kdisp_gfx_text_bounds.

        That function is a pure wrapper over the bbox in the firmware too, and for
        the reason wrappers usually exist here: the two used to be separate walks of
        the same display list, so every op added to one had to be remembered in the
        other.
        """
        xmn, xmx, _ymn, _ymx = self.bbox(cps)
        return xmn, xmx


    def bbox(self, cps):
        """Full ink box relative to the draw origin (x from origin x, y from the
        baseline) — like kdisp_gfx_text_bbox. Mirrors draw()'s cursor rules,
        including the per-font yAdvance shift, so a caller can clamp a draw
        position to keep the glyph on the panel."""
        x = y = 0
        xmn = ymn = 127
        xmx = ymx = -128
        small = mid = False
        for is_op, cp, args in _walk_ops(cps):
            if cp == 0x05: y += 2; continue
            if cp == 0x06: x += 2; continue
            if cp == 0x18: x = y = 0; continue
            if cp == 0x08: x = x - 2 if x > 1 else 0; continue
            if cp == 0x0c: y = y - 2 if y > 1 else 0; continue
            if cp == 0x09: x += (_trunc_div(x, 36) + 1) * 36; continue
            if cp == 0x0a: y += self.base_yadv; x = 0; continue
            if cp == 0x0b: y += (_trunc_div(y, 15) + 1) * 15; continue
            if cp == 0x0d: x = 0; continue
            if cp == HINT_SMALL: small = True; continue
            if cp == HINT_MID: mid = True; continue
            # ⚠️ The composite ops are NOT measured, matching the firmware's own
            # RELATIVE bbox form (`bbox_walk` with resolve=false): MOVE names an
            # ABSOLUTE buffer position, and BADGE/ROT plot AT the cursor through
            # primitives of their own — none of it is knowable without the draw
            # origin the relative form does not have. They are skipped here and
            # drawn in draw(), which does have it, so a MOVE'd mark falls outside
            # the box this reports exactly as it does on the keyboard.
            if is_op and cp in OP_ARGS: continue
            pool, base_yadv = self._pool(cp, mid)
            f = self._font_in(pool, cp); ch = cp
            if f is None:
                # Nothing covers it. The full-size writer substitutes '!' from
                # pool[0] and advances, so the box must too - but the SMALL draw
                # (kdisp_write_gfx_char_half) returns 0: no ink and NO advance.
                # Substituting in a SMALL run would invent both, shifting every
                # following glyph. Skip it, as font_lookup.c does since qmk#252.
                if small: continue
                f = pool[0]; ch = ord('!')
            if not (f.first <= ch <= f.last): continue
            g = f.glyphs[ch - f.first]
            gyadj = f.yAdvance - base_yadv
            w, h = g['width'], g['height']
            if w > 0 and h > 0:
                # In a SMALL run mirror kdisp_write_gfx_char_half — halved extents and
                # offsets, floored — so the measured box still matches the pixels.
                gx = _half_floor(g['xOffset']) if small else g['xOffset']
                gw = (w + 1) // 2 if small else w
                gh = (h + 1) // 2 if small else h
                gy = _half_floor(gyadj + g['yOffset']) if small else gyadj + g['yOffset']
                xmn = min(xmn, x + gx)
                xmx = max(xmx, x + gx + gw - 1)
                ymn = min(ymn, y + gy)
                ymx = max(ymx, y + gy + gh - 1)
            x += (g['xAdvance'] + 1) // 2 if small else g['xAdvance']
        if xmx < xmn: return 0, 0, 0, 0
        return xmn, xmx, ymn, ymx


    def relocate(self, cps, size):
        """glyph_size_remap(): the legend at the requested size, or None to fall
        back to the small face. ALL-OR-NOTHING — a partial hit would mix two faces
        (and so two baselines) in one legend."""
        if size == 0 or not cps:
            return None
        base = GLYPH_SIZE_BASE[size]
        out = []
        for cp in cps:
            if cp < 0x20:
                # A display-list op, not a glyph. The five zero-argument cursor
                # nudges are DROPPED (they were hand-tuned for the small face, and
                # kdisp_gfx_text_bbox cannot see what they do to the draw, so
                # carrying one clips the accent); every other op bails, since the
                # arg-taking ones would have their arguments relocated as glyphs.
                # Mirrors the switch in glyph_size_remap().
                if cp in (0x05, 0x06, 0x08, 0x0B, 0x0C):
                    continue
                return None
            if len(out) + 1 >= GLYPH_SIZE_MAX_LEN + 1:
                return None
            rel = base + cp
            if self._font(rel) is None:
                return None
            out.append(rel)
        return out or None

    def draw(self, setpix, cps, x, y, clearpix=None):
        """Draw a display list. `clearpix` enables HINT_ERASE (\\x14).

        ⚠️ Without a `clearpix` the erase ops plot NOTHING rather than plotting ink:
        an engaged lock badge punches its arrow back out of the solid fill, and
        drawing that as ink would fill the hole in instead of cutting it. Silently
        wrong beats missing here, so it degrades to missing.
        """
        xc, yc = x, y
        erase = False

        def plot(bx, by):
            """Buffer coords -> the caller's pixel space. Mirrors kdisp_plot_ink:
            ONE choke point that every op and the glyph path go through, so the
            erase mode covers the composite ops too. It used to cover the text
            paths only, and a HINT_ERASE before a HALF/BADGE silently drew it lit."""
            vx, vy = bx - BUFFER_X, by
            if not (-OVERSHOOT <= vx < OLED_W + OVERSHOOT
                    and -OVERSHOOT <= vy < OLED_H + OVERSHOOT):
                return
            if erase:
                if clearpix is not None:
                    clearpix(vx, vy)
            else:
                setpix(vx, vy)
        # HINT_SMALL / HINT_MID latch for the REST of the run — there is no "back to
        # full size" op, and \x18 (reset) does not clear them either; it resets the
        # cursor only. The two compose: \x10 after \x16 half-scales the 19px face.
        small = mid = False
        for is_op, cp, args in _walk_ops(cps):
            if cp == 0x05: yc += 2; continue
            if cp == 0x06: xc += 2; continue
            if cp == 0x18: xc, yc = x, y; continue
            if cp == 0x08: xc = xc - 2 if xc > 1 else 0; continue
            if cp == 0x0c: yc = yc - 2 if yc > 1 else 0; continue
            if cp == 0x09: xc += (_trunc_div(xc - x, 36) + 1) * 36; continue
            if cp == 0x0a: yc += self.base_yadv; xc = x; continue
            if cp == 0x0b: yc += (_trunc_div(yc - y, 15) + 1) * 15; continue
            if cp == 0x0d: xc = x; continue
            if cp == HINT_SMALL: small = True; continue
            if cp == HINT_MID: mid = True; continue
            # ⚠️ MOVE is an ABSOLUTE buffer position, so it is an assignment, not an
            # offset -- and the firmware re-applies its jitter offset here, which is
            # 0 in a preview because nothing jitters a static render.
            if cp == 0x0E:
                if args: xc, yc = _int8(args[0]), _int8(args[1])
                continue
            if cp == 0x13:
                if args:
                    # style 2 = solid (engaged); anything else strokes the released
                    # ring. The radius is FIXED, not an argument: the whole point is
                    # to match the baked ICON_CAPSLOCK_* corners.
                    draw_badge_rect(plot, xc, yc, _int8(args[0]), _int8(args[1]),
                                    KDISP_BADGE_RADIUS,
                                    0 if args[2] == 2 else KDISP_BADGE_BORDER)
                continue
            if cp == 0x14: erase = True; continue
            if cp == 0x15:
                if args: self._draw_glyph_rot_half(plot, xc, yc, args[1], args[0])
                continue
            # An op still needing a primitive this model lacks draws NOTHING rather
            # than falling through to the glyph path, where it (and each argument)
            # would render a substituted '!'. unsupported_ops() names them.
            if is_op and cp in OP_ARGS: continue
            pool, base_yadv = self._pool(cp, mid)
            f = self._font_in(pool, cp); ch = cp
            if f is None:
                # ⚠️ Only the FULL-SIZE writer substitutes '!'. kdisp_write_gfx_char_half
                # returns 0 for a glyph it cannot find: it draws nothing and does not
                # advance, which is why a small legend on a keyboard with no font pack
                # comes up blank rather than as a row of '!'.
                if small: continue
                f = pool[0]; ch = ord('!')
            if not (f.first <= ch <= f.last): continue
            g = f.glyphs[ch - f.first]
            gyadj = f.yAdvance - base_yadv
            bo = g['bitmapOffset']
            w, h = g['width'], g['height']
            cb = (h + 7) >> 3             # column-native (OLED page) bytes per column
            if small:
                # 2x2-OR downsample, the same one kdisp_write_gfx_char_half does: plain
                # decimation drops the thin strokes these faces are made of.
                gx0 = xc + _half_floor(g['xOffset'])
                gy0 = yc + _half_floor(gyadj + g['yOffset'])
                for dy in range((h + 1) // 2):
                    for dx in range((w + 1) // 2):
                        lit = False
                        for oy in range(2):
                            for ox in range(2):
                                sx, sy = dx * 2 + ox, dy * 2 + oy
                                if sx >= w or sy >= h: continue
                                if f.bitmap[bo + sx * cb + (sy >> 3)] & (1 << (sy & 7)):
                                    lit = True; break
                            if lit: break
                        if not lit: continue
                        plot(gx0 + dx, gy0 + dy)
                xc += (g['xAdvance'] + 1) // 2
                continue
            gy = yc + gyadj
            for xx in range(w):
                col = bo + xx * cb
                for yy in range(h):
                    if f.bitmap[col + (yy >> 3)] & (1 << (yy & 7)):
                        # ⚠️ Through plot(), NOT setpix, so HINT_ERASE reaches the TEXT
                        # too. The firmware routes every path through kdisp_plot_ink for
                        # exactly this reason, and the C carries the scar: erase used to
                        # cover the text paths only, and an engaged lock badge then drew
                        # its arrow lit instead of punching it out -- a solid blob rather
                        # than an inverted badge.
                        plot(xc + g['xOffset'] + xx, gy + g['yOffset'] + yy)
            xc += g['xAdvance']


def render_key(L: Lang, R: Renderer, lang: str, kc: str, shift: bool, caps: bool,
               channels: bool = False, report: dict | None = None,
               size: int = 0) -> Image.Image:
    """Replicate the per-key draw in keymap.c (process_record_user render path).

    channels=True renders each element into its own colour channel - base->green,
    Shift->blue, AltGr->red - so any overlap between them mixes (base+Shift=cyan,
    base+AltGr=yellow, Shift+AltGr=magenta, all three=white) and is easy to spot.
    """
    li = L.langs.index(lang); row = ROW[kc]
    is_letter = kc[:3] == "KC_" and len(kc) == 4 and kc[3] in LETTERS
    is_num = kc in [f"KC_{d}" for d in "1234567890"]
    cat = "letter" if is_letter else ("num" if is_num else "sym")
    vrow, hrow = SET[cat]

    used_lang, base = L.small(li, row)
    ew, eh = OLED_W + 2 * OVERSHOOT, OLED_H + 2 * OVERSHOOT
    img = Image.new('RGB', (ew, eh), (0, 0, 0)) if channels else Image.new('L', (ew, eh), 0)
    px = img.load()
    if base is None:
        return img

    # base / shift selection mirrors translate_keycode for the requested view
    if caps:
        capc = L.var(used_lang, row, VAR_CAPS)
        if capc is not None and not shift: base = capc
        elif capc is None: shift = not shift
    if shift:
        up = L.var(used_lang, row, VAR_SHIFT)
        if up is not None: base = up

    h_small = get_setting(L, hrow, li, VAR_SMALL); v_small = get_setting(L, vrow, li, VAR_SMALL)
    base_x = 28 + h_small; base_v = v_small

    # --- legend size: plan_main_legend() in poly_keymap.c. At size 0 this is the
    # placement the keyboard has always used; above it the base legend becomes a
    # relocated glyph at the tier's nominal baseline, clamped against its own ink
    # box so a tall accent or deep descender shifts to fit instead of clipping.
    # base_ink is what everything below lays out around.
    # `big_plan` is the (x, y) the relocated glyph is drawn at, bound in the SAME
    # branch that decides there is one — mirroring the firmware's main_legend_t /
    # plan_main_legend() pair. Two separate `if big is not None` blocks with the
    # coordinates living between them read as a use-before-assign to any
    # path-insensitive reader (CodeQL flagged exactly that), and would become one
    # for real the moment anything reassigned `big` in between.
    big_plan = None
    big = R.relocate(base, size)
    if big is not None:
        xmn, xmx, ymn, ymx = R.bbox(big)
        big_x = base_x
        if big_x + xmx > BUFFER_X + SCREEN_WIDTH - 1: big_x = BUFFER_X + SCREEN_WIDTH - 1 - xmx
        if big_x + xmn < BUFFER_X:                    big_x = BUFFER_X - xmn
        big_y = GLYPH_SIZE_BASELINE[size]
        if big_y + ymn < 0:          big_y = -ymn
        if big_y + ymx > OLED_H - 1: big_y = OLED_H - 1 - ymx
        big_plan = (big_x, big_y)
        base_ink_max = big_x + xmx
    else:
        _bmn, _bmx = R.bounds(base)
        base_ink_max = base_x + _bmx

    shift_letter = None; preview_x = preview_v = 0
    if not shift and not caps:
        v_pv = get_setting(L, vrow, li, VAR_SHIFT); h_pv = get_setting(L, hrow, li, VAR_SHIFT)
        if v_pv != HIDE and h_pv != HIDE:
            # mirror translate_keycode_only_shift(): the language's OWN Shift glyph,
            # falling back to the en-US Shift only when the key has neither its own
            # Shift nor its own base. A key that inherits only the base (e.g. the
            # ck-US number keys: en-US digit base + a Cherokee Shift syllable) keeps
            # its Shift - so this is NOT tied to small()'s used_lang (which drops to 0).
            shift_letter = L.var(li, row, VAR_SHIFT)
            if shift_letter is None and L.cell(li, row, VAR_SMALL) is None:
                shift_letter = L.var(0, row, VAR_SHIFT)
            if shift_letter is not None:
                pmin, pmax, pymn, pymx = R.bbox(shift_letter)
                preview_x = 28 + h_pv
                if preview_x + pmin < base_ink_max + 2: preview_x = base_ink_max + 2 - pmin
                if preview_x + pmax > BUFFER_X + SCREEN_WIDTH - 1: preview_x = (BUFFER_X + SCREEN_WIDTH - 1) - pmax
                preview_v = v_pv
                if preview_x + pmin <= base_ink_max:
                    # Only a SMALL base can be lifted: a big one was already clamped
                    # to the panel, so a 6 px lift pushes its ink off the top.
                    if big is None: base_v -= 6
                    preview_v += 4

    # Resolve the AltGr hint BEFORE anything is drawn, for the same reason the Shift
    # preview above is resolved first: the two hints have to be laid out as a pair.
    alt = None; alt_x = 0; v_off = 0
    if not shift and not caps:
        v_off = get_setting(L, vrow, li, VAR_ALTGR); h_off = get_setting(L, hrow, li, VAR_ALTGR)
        if v_off != HIDE and h_off != HIDE:
            alt = L.var(li, row, VAR_ALTGR)
            if alt is not None:
                amin, amax, aymn, aymx = R.bbox(alt)
                # The AltGr glyph is a HINT -- what this key would type with a
                # modifier nobody is holding -- so on a script whose letters fill the
                # keycap it is drawn at HALF size: subordinate to the base legend, and
                # a full-size script glyph is most of what made the two hints fight
                # over the right-hand side.
                #
                # ⚠️ WHICH layouts is DATA, not a size test, and the measurement is
                # why. The intuition is "Arabic and Indic have very large glyphs", but
                # AltGr ink HEIGHT does not separate them: median 20 px on Arabic
                # letters against 21 px on Latin. What differs is that on those
                # layouts the base and Shift are wide too, so the row reads crowded --
                # a per-LAYOUT judgement no glyph measurement can make. It lives in
                # `{letter.altgrhalf}` in lang_lut.xlsx, one cell per language.
                #
                # ⚠️ The size test that REMAINS is only the mark guard: a glyph that is
                # already tiny is destroyed by halving (a Hebrew nikud is 2x3 px and
                # comes out a dot). THAT threshold is measured -- over the 318 distinct
                # AltGr cells the ink-height histogram has an EMPTY BIN at 8 px, marks
                # below it and letterforms from 9 px up -- and it carries most of the
                # Indic layouts, whose letter AltGr hints are mostly bare combining
                # marks at a median 4 px.
                if (is_letter and get_setting(L, ALTGR_HALF_ROW, li, VAR_ALTGR)
                        and aymx - aymn + 1 > ALTGR_HALF_MIN_INK_H and alt[0] != HINT_SMALL):
                    alt = (HINT_SMALL,) + tuple(alt)
                    amin, amax, aymn, aymx = R.bbox(alt)
                # mirror the firmware's right-edge clamp (keymap.c altgr preview)
                alt_x = 28 + h_off
                # At the small size this mark is kept off the legend by its VERTICAL
                # offset; a big legend fills that height, so there the only separation
                # left is horizontal (keymap.c does the same).
                if big is not None and alt_x + amin < base_ink_max + 2:
                    alt_x = base_ink_max + 2 - amin
                if alt_x + amax > BUFFER_X + SCREEN_WIDTH - 1:
                    alt_x = (BUFFER_X + SCREEN_WIDTH - 1) - amax

    # --- keep the two hints off EACH OTHER ---------------------------------
    # Both sit right of the base -- Shift upper, AltGr lower -- and it is their
    # VERTICAL offsets that hold them apart. True for a narrow Latin pair, false for
    # a tall script: on every ar-* KC_F the Shift tick lands inside the AltGr's 29 px
    # box, and bn-BD KC_D shares 57 px. A per-language offset cannot fix that -- the
    # room left over is decided by the WIDTH of this key's three glyphs, so a single
    # number per language would have to satisfy the worst key and would crush the
    # rest into the base.
    #
    # The base is bottom-left and narrow on exactly these keys, so the free space is
    # between it and the (right-clamped) AltGr: pull the Shift LEFT into that gap,
    # never past the base's own 2 px margin, and never to the right (which could only
    # walk it into the clamp). Where three wide glyphs genuinely do not fit on 72 px
    # the pull still shrinks the overlap rather than removing it.
    if shift_letter is not None and alt is not None:
        sx0, sx1 = preview_x + pmin, preview_x + pmax
        sy0, sy1 = BASELINE + preview_v + pymn, BASELINE + preview_v + pymx
        ax0, ax1 = alt_x + amin, alt_x + amax
        ay0, ay1 = BASELINE + v_off + aymn, BASELINE + v_off + aymx
        if sx0 <= ax1 and ax0 <= sx1 and sy0 <= ay1 and ay0 <= sy1:
            want = ax0 - 2 - pmax
            floor = base_ink_max + 2 - pmin
            if want < floor: want = floor
            if want < preview_x: preview_x = want

    EXP = OVERSHOOT
    # report: per-element pixel sets so callers can flag out-of-bounds (a pixel the
    # firmware would clip at the keycap edge) and overlap (a pixel two elements share).
    rpx = {'base': set(), 'shift': set(), 'altgr': set()} if report is not None else None
    def make_setter(ci, name):    # ci: 0=R (AltGr), 1=G (base), 2=B (Shift)
        def s(vx, vy):
            if channels:
                X, Y = vx + EXP, vy + EXP; c = list(px[X, Y]); c[ci] = 255; px[X, Y] = tuple(c)
            else:
                px[vx + EXP, vy + EXP] = 255
            if rpx is not None: rpx[name].add((vx, vy))
        return s
    sp_base, sp_shift, sp_alt = make_setter(1, 'base'), make_setter(2, 'shift'), make_setter(0, 'altgr')

    if big_plan is not None:
        R.draw(sp_base, big, *big_plan)
    else:
        R.draw(sp_base, base, base_x, BASELINE + base_v)
    if shift_letter is not None:
        R.draw(sp_shift, shift_letter, preview_x, BASELINE + preview_v)
    if alt is not None:
        R.draw(sp_alt, alt, alt_x, BASELINE + v_off)
    if report is not None:
        def _oob(s): return sum(1 for (vx, vy) in s if vx < 0 or vx >= OLED_W or vy < 0 or vy >= OLED_H)
        report['oob'] = {k: _oob(v) for k, v in rpx.items()}
        b, sh, al = rpx['base'], rpx['shift'], rpx['altgr']
        report['overlap'] = len((b & sh) | (b & al) | (sh & al))
        report['overlap_detail'] = {'base^shift': len(b & sh), 'base^altgr': len(b & al), 'shift^altgr': len(sh & al)}
        # Per-element ink boxes (x0,x1,y0,y1), so a caller can measure the CLEAR
        # SPACE between elements rather than only whether they collide. Keyed by the
        # element that actually drew, so a key with an AltGr hint but no Shift hint
        # cannot be mis-attributed (positional tagging gets that wrong).
        report['box'] = {k: (min(x for x, _ in v), max(x for x, _ in v),
                             min(y for _, y in v), max(y for _, y in v))
                         for k, v in rpx.items() if v}
    return img


def warn_key(L: Lang, R: Renderer, lang: str, kc: str) -> list[str]:
    """Render a key off-screen with a wide margin and report any element that draws
    out of bounds (clipped at the keycap edge) or overlaps another element."""
    global OVERSHOOT
    save = OVERSHOOT; OVERSHOOT = 60
    try:
        rep: dict = {}
        render_key(L, R, lang, kc, False, False, channels=True, report=rep)
    finally:
        OVERSHOOT = save
    msgs = []
    for el, n in rep['oob'].items():
        if n: msgs.append(f"OUT-OF-BOUNDS {el} {n}px clipped")
    if rep['overlap']:
        d = ", ".join(f"{k}={v}" for k, v in rep['overlap_detail'].items() if v)
        msgs.append(f"OVERLAP {rep['overlap']}px ({d})")
    return msgs


def oled_to_rgb(img: Image.Image, scale: int) -> Image.Image:
    # img is the expanded (OLED_W+2*OVERSHOOT) x (OLED_H+2*OVERSHOOT) buffer.
    if img.mode == 'RGB':
        # --channels: base=green, Shift=blue, AltGr=red already baked in (overlaps
        # mix). Just tint the unlit overshoot margin gray so the viewport edge shows.
        ew, eh = img.size
        if OVERSHOOT:
            px = img.load()
            for y in range(eh):
                iny = OVERSHOOT <= y < OLED_H + OVERSHOOT
                for x in range(ew):
                    if not (iny and OVERSHOOT <= x < OLED_W + OVERSHOOT) and px[x, y] == (0, 0, 0):
                        px[x, y] = (40, 40, 40)
        return img.resize((ew * scale, eh * scale), Image.NEAREST)
    # img is the expanded grayscale buffer.
    # Central OLED_W x OLED_H = the real viewport (white-on-black, hardware-exact). The
    # OVERSHOOT-px border = pixels the device CLIPS: painted dark red, with any lit
    # pixel there shown YELLOW so a glyph cut off at an edge is impossible to miss.
    if OVERSHOOT == 0:
        big = img.resize((OLED_W * scale, OLED_H * scale), Image.NEAREST)
        return Image.merge('RGB', (big, big, big))
    ew, eh = img.size
    src = img.load()
    rgb = Image.new('RGB', (ew, eh)); dst = rgb.load()
    for y in range(eh):
        iny = OVERSHOOT <= y < OLED_H + OVERSHOOT
        for x in range(ew):
            inside = iny and (OVERSHOOT <= x < OLED_W + OVERSHOOT)
            lit = src[x, y] > 0
            if inside:
                dst[x, y] = (255, 255, 255) if lit else (0, 0, 0)
            else:
                dst[x, y] = (255, 255, 0) if lit else (48, 0, 0)
    return rgb.resize((ew * scale, eh * scale), Image.NEAREST)


def main():
    global OVERSHOOT
    ap = argparse.ArgumentParser()
    ap.add_argument('--lang', required=True, help='e.g. ka-GE, ta-IN, vi-VN')
    ap.add_argument('--key', help='single keycode, e.g. KC_Q (default: contact sheet of all keys)')
    ap.add_argument('--shift', action='store_true', help='render the shifted view')
    ap.add_argument('--caps', action='store_true', help='render with caps lock')
    ap.add_argument('--qmk', default=os.path.join(HOME, 'qmk_firmware'))
    ap.add_argument('--cell-scale', type=int, default=3)
    ap.add_argument('--overshoot', type=int, default=OVERSHOOT,
                    help='px of out-of-viewport render to keep & flag (red margin, yellow pixels); 0 = hardware-exact')
    ap.add_argument('--channels', action='store_true',
                    help='overlap-detect: base=green, Shift=blue, AltGr=red; overlaps mix (cyan/yellow/magenta/white)')
    ap.add_argument('--out', default=None)
    ap.add_argument('--check-bounds', action='store_true',
                    help='audit out-of-bounds + element overlap (this lang, or ALL langs with --lang ALL); no image')
    a = ap.parse_args()
    OVERSHOOT = a.overshoot

    pk = os.path.join(a.qmk, 'keyboards', 'polykybd')
    named = load_named_glyphs(os.path.join(pk, 'lang', 'named_glyphs.h'))
    L = Lang(os.path.join(pk, 'lang', 'lang_lut.xlsx'), named)
    if a.lang != 'ALL' and a.lang not in L.langs: sys.exit(f"unknown lang {a.lang}; have {L.langs}")
    R = load_renderer(os.path.join(pk, 'base', 'fonts'))
    s = a.cell_scale

    if a.check_bounds:
        langs = L.langs if a.lang == 'ALL' else [a.lang]
        total = 0
        for lang in langs:
            hits = []
            for kc in ROW:
                w = warn_key(L, R, lang, kc)
                if w: hits.append(f"  {kc.replace('KC_',''):6s} {'; '.join(w)}")
            if hits:
                print(f"{lang}:"); print("\n".join(hits)); total += len(hits)
        print(f"\n{total} key(s) with out-of-bounds/overlap across {len(langs)} lang(s)")
        return

    if a.key:
        w = warn_key(L, R, a.lang, a.key.upper())
        for m in w: print(f"  ⚠ {a.lang} {a.key.upper()}: {m}", file=sys.stderr)
        img = oled_to_rgb(render_key(L, R, a.lang, a.key.upper(), a.shift, a.caps, a.channels), s)
        out = a.out or os.path.join(HERE, 'out', f'oled_{a.lang}_{a.key}.png')
        os.makedirs(os.path.dirname(out), exist_ok=True); img.save(out)
        print("wrote", out); return

    # contact sheet
    try: font = ImageFont.truetype("DejaVuSans.ttf", 10)
    except Exception: font = ImageFont.load_default()
    ew, eh = OLED_W + 2 * OVERSHOOT, OLED_H + 2 * OVERSHOOT
    cw, ch = ew * s, eh * s
    pad, lab = 6, 12
    cols = max(len(r) for r in SHEET); rows = len(SHEET)
    W = cols * (cw + pad) + pad; H = rows * (ch + lab + pad) + pad + 18
    sheet = Image.new('RGB', (W, H), (32, 32, 32)); d = ImageDraw.Draw(sheet)
    title = f"{a.lang}{'  [shift]' if a.shift else ''}{'  [caps]' if a.caps else ''}"
    if a.channels: title += "   channels: base=green Shift=blue AltGr=red (overlap=mix)"
    d.text((pad, 4), title, font=font, fill=(255, 255, 0))
    for ri, krow in enumerate(SHEET):
        for ci, kc in enumerate(krow):
            x = pad + ci * (cw + pad); y = 18 + pad + ri * (ch + lab + pad)
            d.text((x, y), kc.replace("KC_", ""), font=font, fill=(180, 180, 180))
            cell = oled_to_rgb(render_key(L, R, a.lang, kc, a.shift, a.caps, a.channels), s)
            sheet.paste(cell, (x, y + lab))
            # outline the REAL 72x40 viewport (inset by the overshoot margin) so the
            # red border sits clearly outside it
            vx0 = x + OVERSHOOT * s; vy0 = y + lab + OVERSHOOT * s
            d.rectangle([vx0, vy0, vx0 + OLED_W * s - 1, vy0 + OLED_H * s - 1], outline=(70, 70, 70))
    out = a.out or os.path.join(HERE, 'out', f'oled_{a.lang}.png')
    os.makedirs(os.path.dirname(out), exist_ok=True); sheet.save(out)
    print("wrote", out)


if __name__ == '__main__':
    main()
